"""
Import first user (Bland, David) from Staff Hours (2026).xlsm or from projects.csv
(Allocated Week tab exported as CSV). Sheet/CSV: Date, Contract, Location, Section, Employee, Start, Break, Finish, ...

Run from repo root or code-workspace. Requires: pip install openpyxl supabase

Usage:
  python code-workspace/import_payroll_bland_david.py
  python code-workspace/import_payroll_bland_david.py projects.csv
  python code-workspace/import_payroll_bland_david.py --diagnose projects.csv
  python code-workspace/import_payroll_bland_david.py --minimal   # only core columns (if 42703 persists)
  python code-workspace/import_payroll_bland_david.py --fix-imported   # fix already-imported rows: set large_plant_id/workshop_tasks_id when project short_description matches plant/task
  python code-workspace/import_payroll_bland_david.py --list-employees --week 1   # output JSON list of employees for week (excludes Site 1-20)
  python code-workspace/import_payroll_bland_david.py --week 1 --employees "Name1,Name2"   # import only selected employees for that week; skips duplicates
"""

import csv
import json
import os
import re
import sys
from datetime import datetime, date, time, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from supabase import create_client

# Supabase (use service_role to bypass RLS for import)
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://ifvbajmmjkkuvhigcgad.supabase.co")
SUPABASE_KEY = os.environ.get(
    "SUPABASE_SERVICE_ROLE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImlmdmJham1tamtrdXZoaWdjZ2FkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2NDQzNzU1NiwiZXhwIjoyMDgwMDEzNTU2fQ.9n7QrMbp__ZlIHxVv99Dzs4jjkmPwNSayzyNUTZe1C8",
)
EXCEL_PATH = os.environ.get("STAFF_HOURS_EXCEL", r"W:\Master Files\2026\Staff Hours (2026).xlsm")
PROJECTS_CSV = os.environ.get("PROJECTS_CSV", "projects.csv")
SHEET_NAME = "Allocated Week (1)"
# A2:AG = rows 2 to MAX_ROW, cols 1-33 (openpyxl 1-based)
MIN_ROW = 2
MAX_ROW_DEFAULT = 13
MAX_ROW_IMPORT = 500  # when loading by week for import or list-employees
MIN_COL, MAX_COL = 1, 33

# CSV column headers (must match projects.csv header row)
CSV_HEADERS = [
    "Date", "Contract", "Location", "Section", "Employee", "Start", "Break", "Finish", "Hours",
    "Plant 1", "Plant 2", "Plant 3", "Plant 4", "Plant 5", "Plant 6",
    "Mob 1", "Mob 2", "Mob 3", "Mob 4", "Material", "Quantity",
    "FT", "TH", "DT", "NW FT", "NW TH", "NW DT", "Travel", "On Call", "Misc", "Paperwork", "Eating All.", "Country",
]

# Column indices (0-based for code): Date=0, Contract=1, Location=2, Section=3, Employee=4, Start=5, Break=6, Finish=7, Hours=8,
# Plant1-6=9-14, Mob1-4=15-18, Material=19, Qty=20, then 21-32 (Travel=26, OnCall=27, Misc=28 in 0-based)
COL_DATE, COL_CONTRACT, COL_LOCATION, COL_SECTION, COL_EMPLOYEE = 0, 1, 2, 3, 4
COL_START, COL_BREAK, COL_FINISH, COL_HOURS = 5, 6, 7, 8
COL_PLANT_START, COL_PLANT_END = 9, 14   # 9-14 = Plant 1-6
COL_MOB_START, COL_MOB_END = 15, 18      # 15-18 = Mob 1-4
COL_MATERIAL, COL_QTY = 19, 20
COL_TRAVEL, COL_ON_CALL, COL_MISC = 26, 27, 28


def _cell_value(row: tuple, idx: int) -> Any:
    """Get cell value at 0-based column index (row is 1-based openpyxl row)."""
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    return v


def _is_site_placeholder(employee_val: Any) -> bool:
    """Return True if Employee (col 4) is 'Site 1' through 'Site 20' (ignore these rows)."""
    if employee_val is None:
        return False
    s = str(employee_val).strip()
    if not s:
        return False
    lower = s.lower()
    if not lower.startswith("site "):
        return False
    try:
        n = int(lower[5:].strip())
        return 1 <= n <= 20
    except ValueError:
        return False


# Excel epoch: day 1 = 1900-01-01 (Windows Excel)
_EXCEL_EPOCH = date(1899, 12, 30)  # so that 1 -> 1900-01-01


def _parse_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    # Excel serial number (float)
    if isinstance(v, (int, float)):
        try:
            serial = int(float(v))
            if serial > 0:
                return _EXCEL_EPOCH + timedelta(days=serial)
        except (ValueError, OverflowError):
            pass
        return None
    s = str(v).strip()
    if not s:
        return None
    # DD/MM/YYYY or D/M/YYYY
    m = re.match(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        return date(y, mo, d)
    return None


def _parse_time(v: Any) -> Optional[time]:
    if v is None:
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    # Excel time as fraction of day (e.g. 0.395833 = 09:30)
    if isinstance(v, (int, float)):
        try:
            x = float(v)
            if 0 <= x < 1:
                total_secs = int(round(x * 24 * 3600))
                h = total_secs // 3600
                m = (total_secs % 3600) // 60
                return time(h, m)
        except (ValueError, OverflowError):
            pass
        return None
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        return time(int(m.group(1)), int(m.group(2)))
    return None


def _parse_break_minutes(v: Any) -> int:
    """Parse break duration to total minutes (e.g. 0:30 -> 30, 1:00 -> 60)."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0
    s = str(v).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    try:
        return int(float(s))  # minutes only
    except (ValueError, TypeError):
        return 0


def _parse_hours_to_minutes(v: Any) -> int:
    """Parse Hours column (e.g. 03:00, 3:30) to total minutes. Returns 0 if empty or 00:00."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0
    s = str(v).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    try:
        return int(float(s)) * 60  # decimal hours -> minutes
    except (ValueError, TypeError):
        return 0


def _parse_int(v: Any) -> Optional[int]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _parse_num(v: Any) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _travel_minutes(v: Any) -> int:
    """Travel column may be time (e.g. 0:30) or minutes."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0
    s = str(v).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _to_iso_date(d: date) -> str:
    return d.isoformat()


def _load_rows_from_csv(csv_path: str) -> List[Tuple[Any, ...]]:
    """Load data rows from projects.csv; return list of tuples (same column order as Excel: 0=Date, 1=Contract, ...)."""
    rows = []
    path = Path(csv_path)
    if not path.exists():
        return rows
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return rows
        # Map header name -> index for robustness; we still output 0-based tuple matching COL_* order
        col_count = max(len(CSV_HEADERS), len(header))
        for row in reader:
            if not row:
                continue
            # Pad or trim to match expected column count (at least 33 for Travel/On Call/Misc)
            padded = list(row) + [None] * max(0, 33 - len(row))
            rows.append(tuple(padded[:33]))
    return rows


# Cell AH1 (row 1, column 34) holds the last row number with data for the sheet – use it so all rows are imported
AH1_COL = 34  # openpyxl: A=1, ..., AH=34


def _load_rows_from_excel_week(week_num: int, max_row: int = MAX_ROW_IMPORT) -> List[Tuple[Any, ...]]:
    """Load data rows from Excel sheet 'Allocated Week (N)'. Row count taken from cell AH1 if present."""
    rows: List[Tuple[Any, ...]] = []
    if not Path(EXCEL_PATH).exists():
        return rows
    sheet_name = f"Allocated Week ({week_num})"
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return rows
    ws = wb[sheet_name]
    # AH1 = last row number with data; ensure we read all rows
    try:
        ah1_val = ws.cell(row=1, column=AH1_COL).value
        if ah1_val is not None:
            last_row = int(float(ah1_val))
            if last_row >= MIN_ROW:
                max_row = min(max(last_row, MIN_ROW), 20000)  # cap at 20k
    except (TypeError, ValueError):
        pass
    for row in ws.iter_rows(min_row=MIN_ROW, max_row=max_row, min_col=MIN_COL, max_col=MAX_COL, values_only=True):
        rows.append(tuple(row))
    wb.close()
    return rows


def _to_iso_timestamp(d: date, t: Optional[time], tz_offset_hours: int = 0) -> Optional[str]:
    if d is None or t is None:
        return None
    dt = datetime.combine(d, t)
    if tz_offset_hours:
        from datetime import timedelta
        dt = dt + timedelta(hours=tz_offset_hours)
    return dt.strftime("%Y-%m-%dT%H:%M:%S") + ("Z" if tz_offset_hours == 0 else "")


def _list_employees_for_week(week_num: int) -> None:
    """Output JSON array of unique Employee names for the week, excluding Site 1-20. For Flutter to parse."""
    rows = _load_rows_from_excel_week(week_num)
    if not rows:
        print(json.dumps({"employees": [], "error": "No data or sheet not found"}))
        return
    seen: set = set()
    for row in rows:
        if len(row) <= COL_EMPLOYEE:
            continue
        emp = _cell_value(row, COL_EMPLOYEE)
        if emp is None:
            continue
        name = str(emp).strip()
        if not name or _is_site_placeholder(name):
            continue
        seen.add(name)
    out = sorted(seen)
    print(json.dumps({"employees": out}))


def _fix_imported_project_to_plant(sb: Any) -> None:
    """One-time fix: for time_periods with status='imported' and project_id set, if that project's
    short_description matches a large_plant (plant_no or plant_description) or workshop_tasks.task,
    set large_plant_id or workshop_tasks_id and clear project_id. Matching is exact then case-insensitive."""
    # Load imported periods with project_id
    r = sb.table("time_periods").select("id, project_id").eq("status", "imported").not_.is_("project_id", "null").execute()
    rows = r.data or []
    if not rows:
        print("No imported time_periods with project_id found. Nothing to fix.")
        return
    print(f"Found {len(rows)} imported time_period(s) with project_id set.")
    # Load projects (id -> short_description)
    proj_r = sb.table("projects").select("id, short_description").execute()
    projects_by_id = {str(p["id"]): str(p.get("short_description") or "").strip() for p in (proj_r.data or [])}
    # Section -> large_plant id (exact key); also build lowercase key -> id for case-insensitive
    plant_r = sb.table("large_plant").select("id, plant_no, plant_description").execute()
    section_to_plant: Dict[str, str] = {}
    section_to_plant_lower: Dict[str, str] = {}
    for p in (plant_r.data or []):
        pid = p.get("id")
        if not pid:
            continue
        for key in (str(p.get("plant_no") or "").strip(), str(p.get("plant_description") or "").strip()):
            if key:
                section_to_plant[key] = pid
                section_to_plant_lower[key.lower()] = pid
    # Section -> workshop_tasks id (exact + lowercase)
    wr = sb.table("workshop_tasks").select("id, task").execute()
    section_to_workshop: Dict[str, str] = {}
    section_to_workshop_lower: Dict[str, str] = {}
    for w in (wr.data or []):
        wid = w.get("id")
        key = str(w.get("task") or "").strip()
        if wid and key:
            section_to_workshop[key] = wid
            section_to_workshop_lower[key.lower()] = wid

    # "Fleet No 100" -> match plant_no "100"; strip "Fleet No " prefix (case-insensitive)
    def _plant_id_for_short_desc(short_desc: str) -> Optional[str]:
        s = (short_desc or "").strip()
        if not s:
            return None
        pid = section_to_plant.get(s) or section_to_plant_lower.get(s.lower())
        if pid:
            return pid
        lower = s.lower()
        if lower.startswith("fleet no ") and len(lower) > 9:
            plant_no_key = s[9:].strip()  # after "Fleet No "
            pid = section_to_plant.get(plant_no_key) or section_to_plant_lower.get(plant_no_key.lower())
            if pid:
                return pid
        return None

    updated_plant = 0
    updated_workshop = 0
    short_descs_seen: set = set()
    for row in rows:
        tp_id = row.get("id")
        project_id = row.get("project_id")
        if not tp_id or not project_id:
            continue
        short_desc = projects_by_id.get(str(project_id)) or ""
        short_descs_seen.add(short_desc or "(empty)")
        if not short_desc:
            continue
        plant_id = _plant_id_for_short_desc(short_desc)
        workshop_id = section_to_workshop.get(short_desc) or section_to_workshop_lower.get(short_desc.lower())
        if plant_id:
            sb.table("time_periods").update({"large_plant_id": plant_id, "project_id": None}).eq("id", tp_id).execute()
            updated_plant += 1
        elif workshop_id:
            sb.table("time_periods").update({"workshop_tasks_id": workshop_id, "project_id": None}).eq("id", tp_id).execute()
            updated_workshop += 1
    print(f"Fix-imported: updated {updated_plant} row(s) to large_plant_id, {updated_workshop} row(s) to workshop_tasks_id.")
    if updated_plant == 0 and updated_workshop == 0 and short_descs_seen:
        print("Diagnostic: project short_description(s) on imported rows:", sorted(short_descs_seen))
        plant_keys = sorted(set(section_to_plant.keys()))[:20]
        workshop_keys = sorted(set(section_to_workshop.keys()))[:20]
        print("Diagnostic: sample large_plant keys (plant_no / plant_description):", plant_keys)
        print("Diagnostic: sample workshop_tasks keys (task):", workshop_keys)
        print("(Match requires project short_description to equal one of these. If Section in CSV was different, re-import with the updated script or align short_description in projects.)")


def main() -> None:
    diagnose = "--diagnose" in sys.argv
    minimal_payload = "--minimal" in sys.argv
    fix_imported = "--fix-imported" in sys.argv
    list_employees = "--list-employees" in sys.argv
    week_num: Optional[int] = None
    employees_arg: Optional[str] = None
    i = 1
    while i < len(sys.argv):
        a = sys.argv[i]
        if a == "--week" and i + 1 < len(sys.argv):
            try:
                week_num = int(sys.argv[i + 1])
            except ValueError:
                pass
            i += 2
            continue
        if a == "--employees" and i + 1 < len(sys.argv):
            employees_arg = sys.argv[i + 1]
            i += 2
            continue
        i += 1
    csv_path = None
    for a in sys.argv[1:]:
        if a not in ("--diagnose", "--minimal", "--fix-imported", "--list-employees", "--week", "--employees") and not a.startswith("-") and not a.isdigit():
            csv_path = a
            break
    if csv_path is None and Path(PROJECTS_CSV).exists() and not week_num:
        csv_path = PROJECTS_CSV

    if list_employees and week_num is not None:
        _list_employees_for_week(week_num)
        return

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    if fix_imported:
        _fix_imported_project_to_plant(sb)
        return

    rows: List[Tuple[Any, ...]] = []
    if week_num is not None and Path(EXCEL_PATH).exists():
        rows = _load_rows_from_excel_week(week_num)
        print(f"Loaded {len(rows)} data row(s) from Excel week {week_num}: {EXCEL_PATH}")
    if not rows and csv_path and Path(csv_path).exists():
        rows = _load_rows_from_csv(csv_path)
        print(f"Loaded {len(rows)} data row(s) from CSV: {csv_path}")
    if not rows and Path(EXCEL_PATH).exists() and week_num is None:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
            for row in ws.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW_DEFAULT, min_col=MIN_COL, max_col=MAX_COL, values_only=True):
                rows.append(tuple(row))
            wb.close()
            print(f"Loaded {len(rows)} data row(s) from Excel: {EXCEL_PATH}")
        else:
            print(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
            wb.close()
            return
    if not rows:
        if week_num is not None:
            print(f"No data for week {week_num}. Check Excel path and sheet 'Allocated Week ({week_num})'.")
        elif csv_path or Path(PROJECTS_CSV).exists():
            print(f"No data rows in {csv_path or PROJECTS_CSV}")
        else:
            print(f"File not found: {EXCEL_PATH} (and no {PROJECTS_CSV})")
        return

    # Selected employees filter (from --employees "Name1,Name2" or "Name1|Name2" for names containing commas)
    selected_employees: Optional[set] = None
    if employees_arg:
        sep = "|" if "|" in employees_arg else ","
        selected_employees = {n.strip() for n in employees_arg.split(sep) if n.strip()}

    # Cache: display_name -> user_id (for multi-employee import)
    user_cache: Dict[str, str] = {}
    def resolve_user_id(display_name_raw: Any) -> Optional[str]:
        if display_name_raw is None:
            return None
        name = str(display_name_raw).strip()
        if not name or _is_site_placeholder(name):
            return None
        if name in user_cache:
            return user_cache[name]
        us = sb.table("users_setup").select("user_id").eq("display_name", name).limit(1).execute()
        if us.data and len(us.data) > 0:
            user_cache[name] = us.data[0]["user_id"]
            return user_cache[name]
        return None

    user_id: Optional[str] = None  # used in single-user mode and for existing_keys
    # When --week is set, always resolve user per row (so all 6 users get correct user_id). Single-user only when no --week and no --employees.
    use_multi_employee = week_num is not None or selected_employees is not None
    if use_multi_employee:
        # Pre-resolve all unique employee names that appear in the sheet (for duplicate check and per-row resolution)
        if selected_employees is not None:
            for name in selected_employees:
                resolve_user_id(name)
            print(f"Selected employees: {len(selected_employees)}")
        else:
            # --week set but no --employees: resolve every unique Employee in the sheet
            seen_emp = set()
            for row in rows:
                if len(row) <= COL_EMPLOYEE:
                    continue
                emp = _cell_value(row, COL_EMPLOYEE)
                if emp is None or _is_site_placeholder(emp):
                    continue
                name = str(emp).strip()
                if name and name not in seen_emp:
                    seen_emp.add(name)
                    resolve_user_id(name)
            print(f"Unique employees in sheet: {len(seen_emp)} (all will be imported per row)")
    else:
        # Single-user mode: first row employee only (legacy CSV/single-sheet run without --week)
        display_name = None
        for row in rows:
            if len(row) > COL_EMPLOYEE:
                candidate = _cell_value(row, COL_EMPLOYEE)
                if candidate and str(candidate).strip() and not _is_site_placeholder(candidate):
                    display_name = str(candidate).strip()
                    break
        candidates = [display_name] if display_name else []
        candidates += ["Bland, David", "Blank, David"]
        display_name = None
        for name in candidates:
            if not name:
                continue
            us = sb.table("users_setup").select("user_id").eq("display_name", name).limit(1).execute()
            if us.data and len(us.data) > 0:
                user_id = us.data[0]["user_id"]
                display_name = name
                break
        if not user_id:
            print("User not found in users_setup. Tried:", candidates)
            print("First row Employee (col E):", repr(_cell_value(rows[0], COL_EMPLOYEE)) if rows else "n/a")
            return
        def _single_user_resolve(_: Any) -> Optional[str]:
            return user_id
        resolve_user_id = _single_user_resolve
        print(f"User: {display_name} -> {user_id}")
    print(f"Rows read: {len(rows)}")

    # Prefer projects by short_description; need projects list for lookup
    projects_response = sb.table("projects").select("id, client_name, town, short_description").execute()
    projects: List[Dict] = projects_response.data or []

    # large_plant: by plant_no -> id (for Plant 1-6 / Mob 1-4), and Section -> id (for time_periods.large_plant_id)
    plant_response = sb.table("large_plant").select("id, plant_no, plant_description").execute()
    plant_list = plant_response.data or []
    plant_by_no: Dict[str, str] = {str(p["plant_no"]).strip(): p["id"] for p in plant_list if p.get("plant_no")}
    # Section (col 3) can match plant_no or plant_description -> use for time_periods.large_plant_id
    section_to_plant_id: Dict[str, str] = {}
    for p in plant_list:
        pid = p.get("id")
        if not pid:
            continue
        for key in (str(p.get("plant_no") or "").strip(), str(p.get("plant_description") or "").strip()):
            if key:
                section_to_plant_id[key] = pid

    # workshop_tasks: Section can match task -> time_periods.workshop_tasks_id
    workshop_response = sb.table("workshop_tasks").select("id, task").execute()
    section_to_workshop_id: Dict[str, str] = {}
    for w in (workshop_response.data or []):
        wid = w.get("id")
        key = str(w.get("task") or "").strip()
        if wid and key:
            section_to_workshop_id[key] = wid

    def _dup_key(uid: str, wd: str, st: Any) -> Optional[Tuple[str, str, str]]:
        """Normalize (user_id, work_date, start_time) for duplicate check so DB format matches."""
        if not uid or not wd or not st:
            return None
        w = str(wd)[:10]
        s = str(st).strip()
        if not s:
            return None
        # Normalize to YYYY-MM-DDTHH:MM:SSZ so DB variants (.000Z, +00, etc.) match
        if len(s) >= 19:
            s = s[:19].replace(" ", "T") + "Z"
        elif not s.endswith("Z"):
            s = s + "Z"
        return (str(uid), w, s)

    # Build set of existing (user_id, work_date, start_time) to avoid duplicates
    existing_keys: set = set()
    dates_in_rows = []
    for row in rows:
        if len(row) < 9:
            continue
        d = _parse_date(_cell_value(row, COL_DATE))
        if d:
            dates_in_rows.append(d)
    if dates_in_rows:
        min_date = min(dates_in_rows)
        max_date = max(dates_in_rows)
        user_ids_to_check = list(user_cache.values()) if user_cache else []
        if not use_multi_employee and user_id is not None:
            user_ids_to_check = [user_id]
        if user_ids_to_check:
            try:
                r = sb.table("time_periods").select("user_id, work_date, start_time").gte("work_date", min_date.isoformat()).lte("work_date", max_date.isoformat()).in_("user_id", user_ids_to_check).execute()
                for x in (r.data or []):
                    k = _dup_key(str(x.get("user_id") or ""), str(x.get("work_date") or ""), x.get("start_time"))
                    if k:
                        existing_keys.add(k)
            except Exception:
                pass

    inserted = 0
    skipped_no_date = 0
    skipped_no_work = 0
    skipped_site_placeholder = 0
    skipped_unknown_employee = 0
    skipped_not_selected = 0
    skipped_duplicate = 0
    errors: List[str] = []

    for row_idx, row in enumerate(rows):
        if len(row) < 9:
            continue
        employee_cell = _cell_value(row, COL_EMPLOYEE)
        if _is_site_placeholder(employee_cell):
            skipped_site_placeholder += 1
            continue
        row_user_id = resolve_user_id(employee_cell)
        if row_user_id is None:
            skipped_unknown_employee += 1
            continue
        if selected_employees is not None:
            emp_name = str(employee_cell).strip()
            if emp_name not in selected_employees:
                skipped_not_selected += 1
                continue
        raw_date = _cell_value(row, COL_DATE)
        work_date = _parse_date(raw_date)
        if work_date is None:
            skipped_no_date += 1
            if row_idx < 3:
                print(f"  [Skip row {row_idx + MIN_ROW}] no date parsed from col A: {repr(raw_date)}")
            continue
        # Skip rows with no hours (column 8 "Hours" empty or 00:00) – only import rows with worked time
        hours_min = _parse_hours_to_minutes(_cell_value(row, COL_HOURS))
        if hours_min <= 0:
            skipped_no_work += 1
            if diagnose:
                print(f"  Row {row_idx + 2}: {work_date} skip (no hours in column 8: {repr(_cell_value(row, COL_HOURS))})")
            continue
        start_t = _parse_time(_cell_value(row, COL_START))
        finish_t = _parse_time(_cell_value(row, COL_FINISH))
        contract = _cell_value(row, COL_CONTRACT)
        location = _cell_value(row, COL_LOCATION)
        section = _cell_value(row, COL_SECTION)
        break_min = _parse_break_minutes(_cell_value(row, COL_BREAK))
        material = _cell_value(row, COL_MATERIAL)
        qty = _parse_num(_cell_value(row, COL_QTY))
        travel_min = _travel_minutes(_cell_value(row, COL_TRAVEL)) if len(row) > COL_TRAVEL else 0
        on_call_val = _cell_value(row, COL_ON_CALL) if len(row) > COL_ON_CALL else None
        misc_min = _parse_int(_cell_value(row, COL_MISC)) if len(row) > COL_MISC else 0
        if misc_min is None:
            misc_min = 0

        # Resolve Section (col 3): large_plant first (including "Fleet No X" -> plant_no), then workshop_tasks, then project
        section_str = str(section).strip() if section else ""
        large_plant_id = section_to_plant_id.get(section_str) if section_str else None
        if not large_plant_id and section_str and section_str.lower().startswith("fleet no ") and len(section_str) > 9:
            plant_no_key = section_str[9:].strip()
            large_plant_id = section_to_plant_id.get(plant_no_key) or (plant_by_no.get(plant_no_key) if plant_no_key in plant_by_no else None)
        workshop_tasks_id = section_to_workshop_id.get(section_str) if section_str else None
        project_id = None
        if not large_plant_id and not workshop_tasks_id:
            for p in projects:
                if section_str and str(p.get("short_description") or "").strip() == section_str:
                    project_id = p["id"]
                    break
            if project_id is None and (contract or location or section):
                for p in projects:
                    if (str(p.get("client_name") or "").strip() == str(contract or "").strip() and
                        str(p.get("town") or "").strip() == str(location or "").strip() and
                        str(p.get("short_description") or "").strip() == section_str):
                        project_id = p["id"]
                        break

        # Build start_time / finish_time (UTC, date + time)
        start_time_iso = None
        finish_time_iso = None
        if start_t:
            start_time_iso = datetime.combine(work_date, start_t).strftime("%Y-%m-%dT%H:%M:%S.000Z")
        if finish_t:
            finish_time_iso = datetime.combine(work_date, finish_t).strftime("%Y-%m-%dT%H:%M:%S.000Z")

        on_call = False
        if on_call_val is not None:
            if isinstance(on_call_val, bool):
                on_call = on_call_val
            elif isinstance(on_call_val, (int, float)):
                on_call = bool(on_call_val)
            else:
                on_call = str(on_call_val).strip().upper() in ("1", "YES", "TRUE", "Y")

        # Skip if already imported (avoid duplicates)
        work_date_str = _to_iso_date(work_date)
        dup_k = _dup_key(row_user_id, work_date_str, start_time_iso)
        if dup_k and dup_k in existing_keys:
            skipped_duplicate += 1
            continue

        # Build payload. Use --minimal if you get 42703 (undefined column) to send only core columns.
        payload: Dict[str, Any] = {
            "user_id": row_user_id,
            "work_date": work_date_str,
            "start_time": start_time_iso,
            "finish_time": finish_time_iso,
            "status": "imported",
        }
        if not minimal_payload:
            payload["submitted_by"] = row_user_id
            payload["submitted_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
            payload["travel_to_site_min"] = travel_min
            payload["travel_from_site_min"] = 0
            payload["on_call"] = on_call
            payload["misc_allowance_min"] = misc_min
            payload["revision_number"] = 0
        if large_plant_id:
            payload["large_plant_id"] = large_plant_id
        elif workshop_tasks_id:
            payload["workshop_tasks_id"] = workshop_tasks_id
        elif project_id:
            payload["project_id"] = project_id
        if material and not minimal_payload:
            payload["concrete_mix_type"] = str(material)
        if qty is not None and not minimal_payload:
            payload["concrete_qty"] = qty

        if diagnose:
            dest = f"project_id={project_id}" if project_id else (f"large_plant_id={large_plant_id}" if large_plant_id else f"workshop_tasks_id={workshop_tasks_id}")
            if not (project_id or large_plant_id or workshop_tasks_id):
                dest = "no match (project/plant/task)"
            print(f"  Row {row_idx + 2}: {work_date} {start_t}-{finish_t} | {contract or '-'} / {section or '-'} -> {dest} | would insert")
            inserted += 1
            continue

        try:
            ins = sb.table("time_periods").insert(payload).execute()
            if not ins.data or len(ins.data) == 0:
                errors.append(f"Row {row_idx + MIN_ROW}: insert returned no data")
                continue
            tp_id = ins.data[0]["id"]
            inserted += 1
            # Prevent same run from inserting duplicate rows if sheet has repeated rows
            if dup_k:
                existing_keys.add(dup_k)

            # Breaks: 15-30 min = one break at 13:00 or nearest; 45-60 = two (larger at 13:00). Round to 15 min.
            if break_min > 0:
                break_min_15 = round(break_min / 15) * 15
                if break_min_15 <= 30:
                    # One break: 13:00 for break_min_15 minutes (or at period end if period doesn't include 13:00)
                    break_start = datetime.combine(work_date, time(13, 0))
                    break_finish = break_start + timedelta(minutes=break_min_15)
                    if start_t and finish_t:
                        period_start = datetime.combine(work_date, start_t)
                        period_end = datetime.combine(work_date, finish_t)
                        if break_start < period_start or break_start > period_end:
                            # Place at end of period, rounded to 15 min
                            end_rounded = (period_end.minute // 15) * 15
                            break_finish = period_end.replace(minute=end_rounded, second=0, microsecond=0)
                            break_start = break_finish - timedelta(minutes=break_min_15)
                    sb.table("time_period_breaks").insert({
                        "time_period_id": tp_id,
                        "break_start": break_start.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                        "break_finish": break_finish.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                        "display_order": 0,
                    }).execute()
                else:
                    # Two breaks: larger at 13:00
                    b1 = (break_min_15 + 1) // 2
                    b2 = break_min_15 - b1
                    if b1 < b2:
                        b1, b2 = b2, b1
                    for i, mins in enumerate([b2, b1]):
                        start_br = datetime.combine(work_date, time(10 if i == 0 else 13, 0))
                        end_br = start_br + timedelta(minutes=mins)
                        sb.table("time_period_breaks").insert({
                            "time_period_id": tp_id,
                            "break_start": start_br.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                            "break_finish": end_br.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                            "display_order": i,
                        }).execute()

            # Used fleet (Plant 1-6): cols 9-14
            for i in range(COL_PLANT_START, COL_PLANT_END + 1):
                if i >= len(row):
                    break
                val = _cell_value(row, i)
                if val is None:
                    continue
                plant_no = str(val).strip()
                if not plant_no:
                    continue
                pid = plant_by_no.get(plant_no)
                if pid:
                    sb.table("time_period_used_fleet").insert({
                        "time_period_id": tp_id,
                        "large_plant_id": pid,
                        "display_order": i - COL_PLANT_START,
                    }).execute()

            # Mobilised fleet (Mob 1-4): cols 15-18 (col 18 can be concrete_ticket_no if numeric 4+ digits and not plant)
            for i in range(COL_MOB_START, COL_MOB_END + 1):
                if i >= len(row):
                    break
                val = _cell_value(row, i)
                if val is None:
                    continue
                plant_no = str(val).strip()
                if not plant_no:
                    continue
                pid = plant_by_no.get(plant_no)
                if pid:
                    sb.table("time_period_mobilised_fleet").insert({
                        "time_period_id": tp_id,
                        "large_plant_id": pid,
                        "display_order": i - COL_MOB_START,
                    }).execute()
                # Col 18 (index 18): if numeric >= 4 digits and not in plant_no -> concrete_ticket_no (already in payload if we wanted)

        except Exception as e:
            err_msg = str(e)
            # PostgREST 42703 returns JSON with "message" naming the missing column; try to get it
            if hasattr(e, "details") and getattr(e, "details", None):
                err_msg += f" | details: {e.details}"
            if hasattr(e, "message") and getattr(e, "message", None):
                err_msg += f" | message: {e.message}"
            if hasattr(e, "response") and e.response is not None:
                try:
                    body = getattr(e.response, "text", None) or getattr(e.response, "body", None)
                    if body:
                        err_msg += f" | body: {body}"
                except Exception:
                    pass
            errors.append(f"Row {row_idx + 2} ({work_date}): {err_msg}")
            print(f"  API error (row {row_idx + 2}): {err_msg}")

    if skipped_no_date:
        print(f"Skipped {skipped_no_date} row(s) with no parseable date.")
    if skipped_no_work:
        print(f"Skipped {skipped_no_work} row(s) with no hours (column 8 empty or 00:00).")
    if skipped_site_placeholder:
        print(f"Skipped {skipped_site_placeholder} row(s) (Employee is Site 1–20).")
    if skipped_unknown_employee:
        print(f"Skipped {skipped_unknown_employee} row(s) (employee not in users_setup).")
    if skipped_not_selected:
        print(f"Skipped {skipped_not_selected} row(s) (employee not in selected list).")
    if skipped_duplicate:
        print(f"Skipped {skipped_duplicate} row(s) (already imported).")
    print(f"Inserted {inserted} time period(s)." + (" (diagnose: no DB write)" if diagnose else ""))
    if errors:
        print("Errors:")
        for e in errors:
            print("  ", e)
        if any("clocking_distance" in str(e) for e in errors):
            print("\nIf errors mention 'clocking_distance', apply the migration that fixes the trigger:")
            print("  supabase/migrations/20260207180000_fix_clocking_distance_trigger.sql")
            print("Then re-run this script.")


if __name__ == "__main__":
    main()
